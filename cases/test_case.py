import os

from openpyxl.styles import PatternFill,Font

from keywords.key_word_demo import KeyDemo
from selenium import webdriver
import pytest
from common.excel_wr import ParseExcel

# pytestmark =pytest.mark.skip()

def test_01_search():
    wb=ParseExcel('../data/测试数据.xlsx')
    sheet=wb.get_sheet_by_name('Sheet')
    for value in sheet.values:
        args = {}
        args['method'] = value[1]
        args['name'] = value[2]
        args['value'] =value[3]
        args['text']=value[4]
        args['expect'] = value[6]
        # 判断是否为测试用例
        if type(value[0]) is int:
            # 判断是否实例化浏览器
            if value[1] == 'browser':
                driver=KeyDemo('Chrome')
            # 断言判断
            elif 'assert' in value[1]:
                status=getattr(driver,value[1])(**args)
                if status:
                    print(status)
                    sheet.cell(row=value[0]+1,column=9).value = 'pass'
                    sheet.cell(row=value[0] + 1, column=9).fill = PatternFill('solid', fgColor='AACF91')  # 设置格式为绿色
                    sheet.cell(row=value[0] + 1, column=9).font = Font(bold=True)  # 格式为加粗
                else:
                    sheet.cell(row=value[0] + 1, column=9).value = 'fail'
                    sheet.cell(row=value[0] + 1, column=9).fill = PatternFill('solid', fgColor='FF0000')  # 设置格式为绿色
                    sheet.cell(row=value[0] + 1, column=9).font = Font(bold=True)  # 格式为加粗
            else:
                getattr(driver,value[1])(**args)
    wb.save_excel()
    wb.close()


if __name__ == '__main__':
    test_01_search()
    # pytest.main(['-s','--pdb','test_case.py::test_01_search','--alluredir','./result_json'])
    # os.system('allure generate ./result/ -o ./report_allure/   --clean')
